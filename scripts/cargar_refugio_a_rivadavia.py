#!/usr/bin/env python3
"""
Carga productos de El Refugio (Excel) en Kiosco Rivadavia (39).

- Catálogo identidad (código/barras/desc) desde Distribuidora El Negro (33), SIN precios/costos.
- Precio unitario del Excel = costo.
- Precio venta = costo * (1 + margen)  [default margen 30%].
- Match por descripción normalizada (+ fuzzy).

Uso (S1):
  export PYTHONPATH=... && set -a && . back/.env && set +a
  back/venv/bin/python scripts/cargar_refugio_a_rivadavia.py --dry-run
  back/venv/bin/python scripts/cargar_refugio_a_rivadavia.py
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from sqlalchemy.orm import selectinload
from sqlmodel import Session, select

from back.database import engine
from back.gestion import modo_especial_manager
from back.modelos import Articulo
from back.schemas.modo_especial_schemas import ProductoModoEspecialCreate, UnidadMedidaEnum

ID_NEGRO = 33
ID_DESTINO = 39  # Kiosco Rivadavia
MARGEN_DEFAULT = 0.30
FUZZY_MIN = 0.88
DEFAULT_EXCEL = Path("/tmp/lista_de_precios_el_refugio_actualizada.xlsx")


@dataclass
class FilaExcel:
    distribuidor: str
    producto: str
    costo: float


@dataclass
class CatalogoNegro:
    codigo_interno: str
    descripcion: str
    barcodes: list[str]
    categorias: list[str]
    unidad: str
    norm: str


def _strip_accents(texto: str) -> str:
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _norm(texto: str) -> str:
    t = _strip_accents((texto or "").upper())
    # quitar pesos / multiplicadores de empaque
    t = re.sub(r"\d+[.,]?\d*\s*(G|GR|GRS|KG|ML|CC|L|LT|LTS|U|UN|UNI|UNID|UNIDAD)S?\b", " ", t)
    t = re.sub(r"\b\d+\s*X\s*\d+(\s*X\s*\d+)?\b", " ", t)
    t = re.sub(r"[^A-Z0-9]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _tokens(texto: str) -> set[str]:
    stop = {"DE", "DEL", "LA", "LAS", "LOS", "EL", "Y", "CON", "SIN", "X", "PARA"}
    return {tok for tok in _norm(texto).split() if len(tok) > 1 and tok not in stop}


def _similitud(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    na, nb = _norm(a), _norm(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    seq = SequenceMatcher(None, na, nb).ratio()
    ta, tb = _tokens(a), _tokens(b)
    if not ta or not tb:
        return seq
    inter = ta & tb
    jacc = len(inter) / len(ta | tb)
    # exigir overlap real de tokens para no matchear marcas cruzadas
    if len(inter) == 0:
        return 0.0
    if len(inter) == 1 and next(iter(inter)) and len(next(iter(inter))) < 5:
        return min(seq, 0.5)
    return max(seq, 0.5 * seq + 0.5 * jacc)


def _leer_excel(ruta: Path) -> list[FilaExcel]:
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active
    filas: list[FilaExcel] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if i == 1:
            continue
        if not row or len(row) < 3:
            continue
        dist, prod, precio = row[0], row[1], row[2]
        if not prod or precio is None:
            continue
        try:
            costo = float(precio)
        except (TypeError, ValueError):
            continue
        if costo < 0:
            continue
        filas.append(
            FilaExcel(
                distribuidor=str(dist or "Sin distribuidor").strip(),
                producto=str(prod).strip(),
                costo=costo,
            )
        )
    return filas


def _leer_categorias(articulo: Articulo) -> list[str]:
    raw = getattr(articulo, "categorias", None)
    if isinstance(raw, list) and raw:
        return [str(c).strip() for c in raw if str(c).strip()]
    if articulo.categoria and articulo.categoria.nombre:
        return [articulo.categoria.nombre]
    return ["General"]


def _cargar_catalogo_negro(db: Session) -> list[CatalogoNegro]:
    arts = db.exec(
        select(Articulo)
        .where(Articulo.id_empresa == ID_NEGRO, Articulo.activo == True)
        .options(selectinload(Articulo.codigos), selectinload(Articulo.categoria))
    ).all()
    out: list[CatalogoNegro] = []
    for a in arts:
        desc = (a.descripcion or "").strip()
        if not desc or not a.codigo_interno:
            continue
        unidad = (a.unidad_venta or "unidad").strip().lower()
        if unidad not in {e.value for e in UnidadMedidaEnum}:
            unidad = "unidad"
        out.append(
            CatalogoNegro(
                codigo_interno=a.codigo_interno.strip(),
                descripcion=desc,
                barcodes=[c.codigo for c in (a.codigos or []) if c.codigo],
                categorias=_leer_categorias(a),
                unidad=unidad,
                norm=_norm(desc),
            )
        )
    return out


def _match(
    fila: FilaExcel, catalogo: list[CatalogoNegro]
) -> tuple[Optional[CatalogoNegro], str, float]:
    target = _norm(fila.producto)
    if not target:
        return None, "sin_desc", 0.0

    exactos = [c for c in catalogo if c.norm == target]
    if len(exactos) == 1:
        return exactos[0], "exacto", 1.0
    if len(exactos) > 1:
        # preferir el de más barcodes
        best = max(exactos, key=lambda c: len(c.barcodes))
        return best, "exacto_ambiguo", 1.0

    mejor: Optional[CatalogoNegro] = None
    mejor_score = 0.0
    for c in catalogo:
        score = _similitud(target, c.norm)
        if score > mejor_score:
            mejor_score = score
            mejor = c
    if mejor and mejor_score >= FUZZY_MIN:
        return mejor, "fuzzy", mejor_score
    return None, "sin_match", mejor_score


def _precio_venta(costo: float, margen: float) -> float:
    return round(costo * (1.0 + margen), 2)


def _codigo_fallback(idx: int) -> str:
    return f"REF-{idx:04d}"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--margen", type=float, default=MARGEN_DEFAULT)
    parser.add_argument("--destino", type=int, default=ID_DESTINO)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--stock", type=float, default=0.0, help="Stock inicial (default 0)")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"No existe Excel: {args.excel}")
        return 1

    filas = _leer_excel(args.excel)
    print(f"Excel: {len(filas)} productos | margen={args.margen:.0%} | destino={args.destino}")

    with Session(engine) as db:
        catalogo = _cargar_catalogo_negro(db)
        print(f"El Negro: {len(catalogo)} artículos (sin usar precios/costos)")

        stats = {"exacto": 0, "exacto_ambiguo": 0, "fuzzy": 0, "sin_match": 0, "creados": 0, "errores": 0}
        pendientes: list[tuple[FilaExcel, Optional[CatalogoNegro], str, float]] = []

        usados_codigos: set[str] = set()
        for fila in filas:
            hit, tipo, score = _match(fila, catalogo)
            stats[tipo] = stats.get(tipo, 0) + 1
            pendientes.append((fila, hit, tipo, score))

        print("\n=== Match ===")
        for k in ("exacto", "exacto_ambiguo", "fuzzy", "sin_match"):
            print(f"  {k}: {stats[k]}")

        print("\n--- Fuzzy sample ---")
        for fila, hit, tipo, score in pendientes:
            if tipo == "fuzzy":
                print(f"  [{score:.2f}] EXCEL={fila.producto[:50]!r} ← NEGRO={hit.descripcion[:50]!r}")

        print("\n--- Sin match (se crean con código REF-xxxx) ---")
        for fila, hit, tipo, score in pendientes:
            if tipo == "sin_match":
                print(f"  [{score:.2f}] {fila.producto[:70]!r} (${fila.costo})")

        if args.dry_run:
            print("\n[dry-run] No se escribe nada.")
            return 0

        idx_fallback = 1
        for fila, hit, tipo, score in pendientes:
            costo = fila.costo
            venta = _precio_venta(costo, args.margen)
            if hit:
                codigo = hit.codigo_interno
                if codigo in usados_codigos:
                    codigo = f"{codigo}-R{idx_fallback}"
                    idx_fallback += 1
                # descripción del Excel (lista del kiosco); identidad/barras de El Negro
                descripcion = fila.producto
                barcodes = hit.barcodes or None
                categorias = [fila.distribuidor] if fila.distribuidor else (hit.categorias or ["General"])
                unidad = hit.unidad
            else:
                codigo = _codigo_fallback(idx_fallback)
                idx_fallback += 1
                descripcion = fila.producto
                barcodes = None
                categorias = [fila.distribuidor or "General"]
                unidad = "unidad"

            usados_codigos.add(codigo)
            try:
                data = ProductoModoEspecialCreate(
                    codigo_interno=codigo,
                    descripcion=descripcion,
                    precio_venta=venta,
                    precio_costo=costo,
                    categorias=categorias,
                    stock=args.stock,
                    barcodes=barcodes,
                    unidad=UnidadMedidaEnum(unidad),
                    tasa_iva=0.21,
                )
                modo_especial_manager.crear_producto(
                    db,
                    args.destino,
                    data,
                    omitir_conflictos_barcode=True,
                    commit=True,
                )
                # margen persistido para futuras actualizaciones de costo
                art = modo_especial_manager._obtener_articulo_por_codigo(db, args.destino, codigo)
                if art:
                    art.margen_ganancia = args.margen
                    db.add(art)
                    db.commit()
                stats["creados"] += 1
            except Exception as exc:
                stats["errores"] += 1
                print(f"  ERROR {codigo}: {exc}")

        print("\n=== Resultado ===")
        print(f"  creados: {stats['creados']}")
        print(f"  errores: {stats['errores']}")
    return 0 if stats["errores"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
